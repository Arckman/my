#!coding:u8

ini=dict()
with open('info') as f:
    # f.readline()
    for l in f.readlines():
        l=l.strip()
        i=l.split(':')
        if i[0]=='文件名':
            ini['filename']=i[1]
        elif i[0]=='纸张大小':
            ini['_size']=i[1]
        elif i[0]=='第一页字体':
            ini['_1_font']=i[1]
        elif i[0]=='第一页字号':
            ini['_1_size']=i[1]
        elif i[0]=='编号位置':
            ini['_po_code']=i[1]
        elif i[0]=='第二页字体':
            ini['_2_font']=i[1]
        elif i[0]=='第二页字号':
            ini['_2_size']=i[1]
        elif i[0]=='姓名位置':
            ini['_po_name']=i[1]
        elif i[0]=='性别位置':
            ini['_po_sex']=i[1]
        elif i[0]=='出生地点位置':
            ini['_po_born']=i[1]
        elif i[0]=='身份证号位置':
            ini['_po_id']=i[1]
        elif i[0]=='专业名称位置':
            ini['_po_category']=i[1]
        elif i[0]=='资格名称位置':
            ini['_po_cert']=i[1]
        elif i[0]=='授予时间位置':
            ini['_po_time']=i[1]
def str2tuple(s):
    t=s.split(',')
    return (int(t[0]),int(t[1]))
ini['_size']=str2tuple(ini['_size'])
ini['_1_size']=int(ini['_1_size'])
ini['_po_code']=str2tuple(ini['_po_code'])
ini['_2_size']=int(ini['_2_size'])
ini['_po_name']=str2tuple(ini['_po_name'])
ini['_po_sex']=str2tuple(ini['_po_sex'])
ini['_po_born']=str2tuple(ini['_po_born'])
ini['_po_id']=str2tuple(ini['_po_id'])
ini['_po_category']=str2tuple(ini['_po_category'])
ini['_po_cert']=str2tuple(ini['_po_cert'])
ini['_po_time']=str2tuple(ini['_po_time'])


from openpyxl import Workbook,load_workbook
from openpyxl.utils.datetime import from_excel
sheetnames=[]
data=[]
import time
import datetime,dateutil
def readData(filename):
    global data,sheetnames
    wb=load_workbook(filename)
    assert isinstance(wb,Workbook)
    sheetnames=wb.sheetnames
    for sheet in sheetnames:
        sheet=wb.get_sheet_by_name(sheet)
        s=[]
        max_row=sheet.max_row
        max_column=sheet.max_column
        for r in range(2,max_row+1):
            p=dict()
            p['name']=sheet.cell(row=r,column=1).value.strip()
            p['sex']=sheet.cell(row=r,column=2).value.strip()
            p['born']=sheet.cell(row=r,column=3).value.strip()
            p['id']=sheet.cell(row=r,column=4).value.strip()
            p['category']=sheet.cell(row=r,column=5).value.strip()
            p['cert']=sheet.cell(row=r,column=6).value.strip()
            p['time']=from_excel(sheet.cell(row=r,column=7).value)
            p['code']=sheet.cell(row=r,column=8).value.strip()
            s.append(p)
        data.append(s)

import os
from PIL import Image, ImageDraw, ImageFont
def drawFirstPage(dir,data,ini):
    # make a blank image for the text, initialized to transparent text color
    txt = Image.new('RGBA',ini['_size'], (255,255,255,0))

    # get a font
    fnt = ImageFont.truetype("SIMSUN.TTC",size=ini['_1_size'])
    # get a drawing context
    d = ImageDraw.Draw(txt)

    # draw text, full opacity
    d.text(ini['_po_code'], data['code'], font=fnt, fill=(0,0,0,255))

    txt.save(dir+os.path.sep+data['name']+'-'+data['id'][-4:]+'-编号.png')

def drawSecondPage(dir,data,ini):
        # make a blank image for the text, initialized to transparent text color
    txt = Image.new('RGBA',ini['_size'], (255,255,255,0))

    # get a font
    fnt = ImageFont.truetype("SIMSUN.TTC",size=ini['_2_size'])
    # get a drawing context
    d = ImageDraw.Draw(txt)

    # draw text, full opacity
    d.text(ini['_po_name'], data['name'], font=fnt, fill=(0,0,0,255))
    d.text(ini['_po_sex'], data['sex'], font=fnt, fill=(0,0,0,255))
    d.text(ini['_po_born'], data['born'], font=fnt, fill=(0,0,0,255))
    d.text(ini['_po_id'], data['id'], font=fnt, fill=(0,0,0,255))
    d.text(ini['_po_category'], data['category'], font=fnt, fill=(0,0,0,255))
    d.text(ini['_po_cert'], data['cert'], font=fnt, fill=(0,0,0,255))
    date=data['time'].strftime('%Y年%m月%d日')#2018年03月08日
    # print(date[-3])
    if date[5]=='0':
        date=date[:5]+date[6:]
    if date[-3]=='0':
        date=date[:-3]+date[-2:]
    d.text(ini['_po_time'], date, font=fnt, fill=(0,0,0,255))
    txt.save(dir+os.path.sep+data['name']+'-'+data['id'][-4:]+'-信息.png')

def drawPics():
    global sheetnames,data,ini
    for i in range(len(sheetnames)):
    # for i in range(1,2):
        sheetname=sheetnames[i]
        path='.'+os.path.sep+sheetname
        if not os.path.exists(path):
            os.mkdir(path)
        for j in data[i]:
            drawFirstPage(path,j,ini)
            drawSecondPage(path,j,ini)

readData(ini['filename'])
drawPics()