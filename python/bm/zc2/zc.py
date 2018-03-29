#!coding:u8

ini = dict()


def readInfo():
    global ini
    with open('info',encoding='u8') as f:   #fix reading bug in windows
        # f.readline()
        for l in f.readlines():
            l = l.strip()
            i = l.split(':')
            if i[0] == '文件名':
                ini['filename'] = i[1]
            elif i[0] == '纸张大小':
                ini['_size'] = i[1]
            elif i[0] == '第一页字体':
                ini['_1_font'] = i[1]
            elif i[0] == '第一页字号':
                ini['_1_size'] = i[1]
            elif i[0] == '编号位置':
                ini['_po_code'] = i[1]
            elif i[0] == '第二页字体':
                ini['_2_font'] = i[1]
            elif i[0] == '第二页字号':
                ini['_2_size'] = i[1]
            elif i[0] == '姓名位置':
                ini['_po_name'] = i[1]
            elif i[0] == '性别位置':
                ini['_po_sex'] = i[1]
            elif i[0] == '出生地点位置':
                ini['_po_born'] = i[1]
            elif i[0] == '身份证号位置':
                ini['_po_id'] = i[1]
            elif i[0] == '专业名称位置':
                ini['_po_category'] = i[1]
            elif i[0] == '资格名称位置':
                ini['_po_cert'] = i[1]
            elif i[0] == '授予时间位置':
                ini['_po_time'] = i[1]
            elif i[0]=='测试':
                ini['_debug']=i[1].lower()

    def str2tuple(s):
        t = s.split(',')
        return (float(t[0]), float(t[1]))

    ini['_size'] = str2tuple(ini['_size'])
    ini['_1_size'] = int(ini['_1_size'])
    ini['_po_code'] = str2tuple(ini['_po_code'])
    ini['_2_size'] = int(ini['_2_size'])
    ini['_po_name'] = str2tuple(ini['_po_name'])
    ini['_po_sex'] = str2tuple(ini['_po_sex'])
    ini['_po_born'] = str2tuple(ini['_po_born'])
    ini['_po_id'] = str2tuple(ini['_po_id'])
    ini['_po_category'] = str2tuple(ini['_po_category'])
    ini['_po_cert'] = str2tuple(ini['_po_cert'])
    ini['_po_time'] = str2tuple(ini['_po_time'])
    ini['_debug']=True if ini['_debug']=='true' else False

from openpyxl import Workbook, load_workbook
from openpyxl.utils.datetime import from_excel
sheetnames = []
data = []
import time
import datetime


def readData(filename):
    global data, sheetnames
    wb = load_workbook(filename)
    assert isinstance(wb, Workbook)
    sheetnames = wb.sheetnames
    for sheet in sheetnames:
        sheet = wb.get_sheet_by_name(sheet)
        s = []
        max_row = sheet.max_row
        max_column = sheet.max_column
        for r in range(2, max_row+1):
            p = dict()
            p['name'] = sheet.cell(row=r, column=1).value.strip()
            p['sex'] = sheet.cell(row=r, column=2).value.strip()
            p['born'] = sheet.cell(row=r, column=3).value.strip()
            p['id'] = sheet.cell(row=r, column=4).value.strip()
            p['category'] = sheet.cell(row=r, column=5).value.strip()
            p['cert'] = sheet.cell(row=r, column=6).value.strip()
            tt = from_excel(sheet.cell(row=r, column=7).value)
            # tt = tt.strftime('%Y年%m月%d日')  # 2018年03月08日
            tt=tt.strftime('%Y{0}%m{1}%d{2}').format(*'年月日')
            #fix bug for strftime() in windows ,referring to https://stackoverflow.com/questions/16034060/python3-datetime-datetime-strftime-failed-to-accept-utf-8-string-format
            # print(tt[-3])
            if tt[5] == '0':
                tt = tt[:5]+tt[6:]
            if tt[-3] == '0':
                tt = tt[:-3]+tt[-2:]
            p['time'] = tt
            p['code'] = sheet.cell(row=r, column=8).value.strip()
            s.append(p)
        data.append(s)


from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from reportlab.lib.colors import red,pink
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import os
# cm=180/2.54
# cm=cm/14.5*17.3

def drawFristPdf(dir, data, ini):
    c = canvas.Canvas(dir+os.path.sep+data['name']+'-'+data['id'][-4:] +
                      '-1-编号.pdf', pagesize=(ini['_size'][0]*cm, ini['_size'][1]*cm))
    c.setFont('song',ini['_1_size'])
    c.drawString(ini['_po_code'][0]*cm, ini['_po_code'][1]*cm, data['code'])
    if ini['_debug']:
        c.setLineWidth(4)
        c.setStrokeColor(red)
        c.rect(0,0,ini['_size'][0]*cm,ini['_size'][1]*cm)
        c.setLineWidth(1)
        c.setStrokeColor(pink)
        c.setDash(4,3)
        p=c.beginPath()
        p.moveTo(0,ini['_size'][1]*cm/4)
        p.lineTo(ini['_size'][0]*cm,ini['_size'][1]*cm/4)
        p.moveTo(0,ini['_size'][1]*cm/2)
        p.lineTo(ini['_size'][0]*cm,ini['_size'][1]*cm/2)
        p.moveTo(0,ini['_size'][1]*cm/4*3)
        p.lineTo(ini['_size'][0]*cm,ini['_size'][1]*cm/4*3)
        p.moveTo(ini['_size'][0]*cm/4,0)
        p.lineTo(ini['_size'][0]*cm/4,ini['_size'][1]*cm)
        p.moveTo(ini['_size'][0]*cm/2,0)
        p.lineTo(ini['_size'][0]*cm/2,ini['_size'][1]*cm)
        p.moveTo(ini['_size'][0]*cm/4*3,0)
        p.lineTo(ini['_size'][0]*cm/4*3,ini['_size'][1]*cm)
        p.close()
        c.drawPath(p)
    c.showPage()
    c.save()


def drawSecondPdf(dir, data, ini):
    c = canvas.Canvas(dir+os.path.sep+data['name']+'-'+data['id'][-4:] +
                      '-2-信息.pdf', pagesize=(ini['_size'][0]*cm, ini['_size'][1]*cm))
    c.setFont('song', ini['_2_size'])
    c.drawCentredString(ini['_po_name'][0]*cm, ini['_po_name'][1]*cm, data['name'])
    c.drawCentredString(ini['_po_sex'][0]*cm, ini['_po_sex'][1]*cm, data['sex'])
    c.drawCentredString(ini['_po_born'][0]*cm, ini['_po_born'][1]*cm, data['born'])
    c.drawCentredString(ini['_po_id'][0]*cm, ini['_po_id'][1]*cm, data['id'])
    c.drawCentredString(ini['_po_category'][0]*cm,
                 ini['_po_category'][1]*cm, data['category'])
    c.drawCentredString(ini['_po_cert'][0]*cm, ini['_po_cert'][1]*cm, data['cert'])
    c.drawCentredString(ini['_po_time'][0]*cm, ini['_po_time'][1]*cm, data['time'])
    if ini['_debug']:
        c.setStrokeColor(red)
        c.setLineWidth(4)
        c.rect(0,0,ini['_size'][0]*cm,ini['_size'][1]*cm)
        c.setLineWidth(1)
        c.setStrokeColor(pink)
        c.setDash(4,3)
        p=c.beginPath()
        p.moveTo(0,ini['_size'][1]*cm/4)
        p.lineTo(ini['_size'][0]*cm,ini['_size'][1]*cm/4)
        p.moveTo(0,ini['_size'][1]*cm/2)
        p.lineTo(ini['_size'][0]*cm,ini['_size'][1]*cm/2)
        p.moveTo(0,ini['_size'][1]*cm/4*3)
        p.lineTo(ini['_size'][0]*cm,ini['_size'][1]*cm/4*3)
        p.moveTo(ini['_size'][0]*cm/4,0)
        p.lineTo(ini['_size'][0]*cm/4,ini['_size'][1]*cm)
        p.moveTo(ini['_size'][0]*cm/2,0)
        p.lineTo(ini['_size'][0]*cm/2,ini['_size'][1]*cm)
        p.moveTo(ini['_size'][0]*cm/4*3,0)
        p.lineTo(ini['_size'][0]*cm/4*3,ini['_size'][1]*cm)
        p.close()
        c.drawPath(p)
    c.showPage()
    c.save()


def drawPdfs():
    global sheetnames, data, ini
    pdfmetrics.registerFont(TTFont('song', ini['_1_font']))
    
    for i in range(len(sheetnames)):
        # for i in range(1,2):
        sheetname = sheetnames[i]
        path = '.'+os.path.sep+sheetname
        if not os.path.exists(path):
            os.mkdir(path)
        for j in data[i]:
            drawFristPdf(path, j, ini)
            drawSecondPdf(path, j, ini)

readInfo()
readData(ini['filename'])
drawPdfs()
